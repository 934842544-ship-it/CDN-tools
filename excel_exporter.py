# -*- coding: utf-8 -*-
"""生成导出对账单 Excel"""
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from db import get_conn
from config import FILTER_ATTRIBUTE_16

EXPORT_COLUMNS = [
    ("customer_name", "客户名称"),
    ("city_company", "市公司"),
    ("account_type", "账户类型"),
    ("email", "客户邮箱"),
    ("product_name", "产品名称"),
    ("occurrence_month", "发生月份"),
    ("billing_month", "列账月"),
    ("usage_volume", "用量"),
    ("billing_amount", "计费列账金额(扣除坏账率)"),
    ("is_settled", "是否销账"),
]

HEADER_FILL = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def export_billing(city, start_month, end_month):
    """按地市+账期范围生成 Excel，返回 (文件名, bytes_io)"""
    conn = get_conn()
    try:
        rows = conn.execute(f"""
            SELECT {','.join(c for c, _ in EXPORT_COLUMNS)}
            FROM customer_billing
            WHERE attribute_16 = ?
              AND city_company = ?
              AND occurrence_month >= ?
              AND occurrence_month <= ?
            ORDER BY occurrence_month ASC, customer_name ASC
        """, (FILTER_ATTRIBUTE_16, city, start_month, end_month)).fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对账单"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXPORT_COLUMNS))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{city} 对账单（{start_month} 至 {end_month}）"
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 32

    for idx, (_, cn_name) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx)
        cell.value = cn_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 26

    total_amount = 0
    total_volume = 0
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, (col, _) in enumerate(EXPORT_COLUMNS, start=1):
            val = row[col]
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

            if col == "is_settled":
                cell.value = "已销账" if val == 1 else "未销账"
                cell.alignment = CENTER
            elif col in ("usage_volume", "billing_amount"):
                cell.value = float(val) if val is not None else 0
                cell.alignment = RIGHT
                cell.number_format = "#,##0.00"
            else:
                cell.value = val if val is not None else ""
                cell.alignment = LEFT if col in ("customer_name", "email", "product_name") else CENTER

        total_amount += float(row["billing_amount"] or 0)
        total_volume += float(row["usage_volume"] or 0)

    last_row = len(rows) + 2
    if last_row >= 2:
        total_row = last_row + 1
        ws.cell(row=total_row, column=1).value = "合计"
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
        for c in range(1, len(EXPORT_COLUMNS) + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.font = Font(name="微软雅黑", size=10, bold=True)
            cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
        vol_cell = ws.cell(row=total_row, column=8)
        vol_cell.value = total_volume
        vol_cell.alignment = RIGHT
        vol_cell.number_format = "#,##0.00"
        amt_cell = ws.cell(row=total_row, column=9)
        amt_cell.value = total_amount
        amt_cell.alignment = RIGHT
        amt_cell.number_format = "#,##0.00"
        ws.row_dimensions[total_row].height = 24

    col_widths = {
        "customer_name": 22,
        "city_company": 12,
        "account_type": 12,
        "email": 26,
        "product_name": 24,
        "occurrence_month": 12,
        "billing_month": 12,
        "usage_volume": 12,
        "billing_amount": 20,
        "is_settled": 10,
    }
    for idx, (col, _) in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col, 14)

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{city}_对账单_{start_month}_{end_month}.xlsx"
    return filename, buf
